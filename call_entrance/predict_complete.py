#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
【完整预测流程】YOLO框检测 + ResNet152油量识别

【工作流程】
1. YOLO检测油表框位置
2. 从原图中裁剪框内区域
3. ResNet152识别框内油量
4. 绘制结果并保存

【使用方法】
python predict_complete.py <image_path_or_dir>

【示例】
python predict_complete.py ./test_images
python predict_complete.py image.jpg
"""
import torch
import torch.nn as nn
from ultralytics import YOLO
from pathlib import Path
import cv2
import numpy as np
from datetime import datetime
import torchvision.models as models
from torchvision.models import ResNet50_Weights
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def expand_box_if_edge_touching(x1, y1, x2, y2, img_h, img_w, expand_ratio=0.1):
    """如果框贴边，自动扩展框以包含更多上下文

    Args:
        x1, y1, x2, y2: 原框坐标
        img_h, img_w: 图片高宽
        expand_ratio: 扩展比例（框大小的比例）

    Returns:
        (x1_new, y1_new, x2_new, y2_new): 扩展后的框坐标
    """

    w = x2 - x1
    h = y2 - y1

    expand_w = int(w * expand_ratio)
    expand_h = int(h * expand_ratio)

    # 检查并扩展各边界
    x1_new = max(0, x1 - expand_w)
    y1_new = max(0, y1 - expand_h)
    x2_new = min(img_w, x2 + expand_w)
    y2_new = min(img_h, y2 + expand_h)

    # 如果有边贴边，再次扩展对面的边
    if x1_new == 0 and x2_new < img_w:
        x2_new = min(img_w, x2_new + expand_w * 2)
    if x2_new == img_w and x1_new > 0:
        x1_new = max(0, x1_new - expand_w * 2)
    if y1_new == 0 and y2_new < img_h:
        y2_new = min(img_h, y2_new + expand_h * 2)
    if y2_new == img_h and y1_new > 0:
        y1_new = max(0, y1_new - expand_h * 2)

    return x1_new, y1_new, x2_new, y2_new


def is_box_valid(x1, y1, x2, y2, img_h, img_w, min_box_size=50, min_box_ratio=0.3, max_box_ratio=3.0):
    """检查框是否有效

    Args:
        x1, y1, x2, y2: 框的坐标
        img_h, img_w: 图片高度和宽度
        min_box_size: 最小框大小（像素）
        min_box_ratio: 最小框纵横比（高/宽）
        max_box_ratio: 最大框纵横比（高/宽，防止过细的框）

    Returns:
        (is_valid, reason) - 是否有效及原因
    """

    w = x2 - x1
    h = y2 - y1

    # 检查基本尺寸
    if w <= 0 or h <= 0:
        return False, "框尺寸为0"

    if w < min_box_size or h < min_box_size:
        return False, f"框太小 ({w}x{h} < {min_box_size}x{min_box_size})"

    # 检查宽度（框不能太窄）
    aspect_ratio = h / w if w > 0 else float('inf')

    if aspect_ratio < min_box_ratio:
        return False, f"框太宽 (比例 {aspect_ratio:.2f} < {min_box_ratio})"

    if aspect_ratio > max_box_ratio:
        return False, f"框太高 (比例 {aspect_ratio:.2f} > {max_box_ratio})"


    return True, "有效框"


class ResNetFuelNet(nn.Module):
    """ResNet50迁移学习油量识别网络"""

    def __init__(self, pretrained=True):
        super().__init__()

        # 加载预训练的ResNet50（使用新的weights参数）
        if pretrained:
            self.backbone = models.resnet50(weights=ResNet50_Weights.DEFAULT)
        else:
            self.backbone = models.resnet50(weights=None)

        num_features = self.backbone.fc.in_features
        self.backbone.fc = nn.Sequential(
            nn.Linear(num_features, 512),
            nn.ReLU(),
            nn.Dropout(0.4),
            nn.Linear(512, 256),
            nn.ReLU(),
            nn.Dropout(0.3),
            nn.Linear(256, 128),
            nn.ReLU(),
            nn.Dropout(0.2),
            nn.Linear(128, 1),
            nn.Sigmoid()
        )

    def forward(self, x):
        return self.backbone(x)


def predict_complete(image_path, result_dir='results_predict', yolo_conf=0.5, min_box_size=50, min_box_ratio=0.3):
    """完整预测流程：YOLO检测框 + ResNet50识别油量

    Args:
        image_path: 图片路径或目录
        result_dir: 结果保存目录（支持相对路径和绝对路径）
        yolo_conf: YOLO置信度阈值（0-1，默认0.6）
        min_box_size: 最小框大小（像素，默认50）
        min_box_ratio: 最小框纵横比（高/宽，默认0.3，范围0.1-1.0）
    """

    print("\n" + "="*70)
    print("【YOLO + ResNet50 完整预测】")
    print("="*70 + "\n")

    # 获取脚本所在目录
    script_dir = Path(__file__).parent.absolute()
    print(f"📍 项目目录: {script_dir}\n")

    # 创建结果目录
    result_dir = Path(result_dir)
    if not result_dir.is_absolute():
        result_dir = script_dir / result_dir
    result_dir.mkdir(exist_ok=True)

    # 创建子目录
    (result_dir / 'detected').mkdir(exist_ok=True)
    (result_dir / 'crops').mkdir(exist_ok=True)

    device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
    print(f"🖥️  使用设备: {device}\n")

    # 打印配置信息
    print(f"⚙️  检测配置:")
    print(f"   YOLO置信度: {yolo_conf:.1%}")
    print(f"   最小框大小: {min_box_size}x{min_box_size} 像素")
    print(f"   最小框比例(高/宽): {min_box_ratio:.1f}\n")

    # ========== 加载YOLO模型 ==========
    print("📦 加载YOLO11m检测模型...\n")

    # yolo_weights = script_dir / 'runs' / 'fuel_yolo' / 'detect' / 'weights' / 'best.pt'
    yolo_weights = script_dir / 'models' / 'using_model' / 'yolo_best.pt' # 第一阶段模型

    if not yolo_weights.exists():
        print(f"❌ 找不到YOLO模型: {yolo_weights}")
        print("   请先运行: python train_yolo_fuel_resnet.py")
        return False

    yolo_model = YOLO(str(yolo_weights))
    print(f"✅ YOLO模型加载成功\n")

    # ========== 加载ResNet模型 ==========
    print("📦 加载ResNet50油量识别模型...\n")

    resnet_path = script_dir / 'models' / 'using_model' / 'fuel_resnet_model.pth' # 第二阶段模型

    if not resnet_path.exists():
        print(f"❌ 找不到ResNet模型: {resnet_path}")
        print("   请先运行: python train_yolo_fuel_resnet.py")
        return False

    resnet_model = ResNetFuelNet(pretrained=False)
    resnet_model.load_state_dict(torch.load(str(resnet_path), map_location=device))
    resnet_model.to(device)
    resnet_model.eval()
    print(f"✅ ResNet模型加载成功\n")

    # ========== 获取测试图片 ==========
    image_path = Path(image_path)

    if image_path.is_dir():
        test_images = sorted(
            list(image_path.glob('*.jpg')) +
            list(image_path.glob('*.jpeg')) +
            list(image_path.glob('*.png'))
        )
        print(f"📁 找到 {len(test_images)} 张图片\n")
    else:
        test_images = [image_path]
        print(f"📁 处理 1 张图片\n")

    if not test_images:
        print("❌ 没有找到图片")
        return False

    # ========== 预测 ==========
    print("🎯 开始预测...\n")

    results_list = []

    for img_path in test_images:
        img_name = img_path.stem
        print(f"📸 {img_path.name}")

        # 读取图片
        img = cv2.imread(str(img_path))
        if img is None:
            print(f"   ❌ 无法读取图片\n")
            continue

        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        h, w = img_rgb.shape[:2]
        img_vis = img_rgb.copy()

        # ========== 第1步：YOLO检测框 ==========
        yolo_results = yolo_model(str(img_path), conf=yolo_conf, verbose=False)

        detected_boxes = []
        fuel_predictions = []

        for result in yolo_results:
            if len(result.boxes) == 0:
                print(f"   ⚠️  未检测到框")
                continue

            # 收集所有检测到的框
            all_boxes = []
            invalid_boxes = []

            for i, box in enumerate(result.boxes):
                x1, y1, x2, y2 = map(int, box.xyxy[0])
                conf = box.conf[0].item()

                # 限制框在图片范围内
                x1, y1, x2, y2 = max(0, x1), max(0, y1), min(w, x2), min(h, y2)

                if x2 <= x1 or y2 <= y1:
                    continue

                # 【检查框是否完整】- 传入图片尺寸
                is_valid, reason = is_box_valid(x1, y1, x2, y2, h, w, min_box_size, min_box_ratio)

                if is_valid:
                    all_boxes.append({
                        'box': (x1, y1, x2, y2),
                        'conf': conf,
                        'index': i
                    })
                else:
                    invalid_boxes.append({
                        'box': (x1, y1, x2, y2),
                        'conf': conf,
                        'reason': reason
                    })

            # 输出被过滤的框
            if invalid_boxes:
                print(f"   ⚠️  过滤掉 {len(invalid_boxes)} 个不完整的框:")
                for invalid in invalid_boxes:
                    x1, y1, x2, y2 = invalid['box']
                    w_box = x2 - x1
                    h_box = y2 - y1
                    print(f"      - 置信度 {invalid['conf']:.1%}, 大小 {w_box}x{h_box}, 原因: {invalid['reason']}")

            if not all_boxes:
                print(f"   ⚠️  未检测到有效框\n")
                continue

            # 只保留置信度最高的框
            best_box = max(all_boxes, key=lambda x: x['conf'])
            x1, y1, x2, y2 = best_box['box']
            conf = best_box['conf']
            box_index = best_box['index']

            if len(all_boxes) > 1:
                print(f"   ℹ️  检测到 {len(all_boxes)} 个完整框，仅保留置信度最高的框")

            # ========== 第1.5步：框边界扩展（如果贴边） ==========
            x1, y1, x2, y2 = best_box['box']

            # 检查是否贴边
            margin_left = x1
            margin_top = y1
            margin_right = w - x2
            margin_bottom = h - y2

            is_edge_touch = (margin_left < 10 or margin_top < 10 or
                           margin_right < 10 or margin_bottom < 10)

            if is_edge_touch:
                x1_exp, y1_exp, x2_exp, y2_exp = expand_box_if_edge_touching(x1, y1, x2, y2, h, w, expand_ratio=0.15)
                print(f"   ℹ️  框贴边，已自动扩展")
                print(f"      原框: ({x1}, {y1}, {x2}, {y2})")
                print(f"      扩展后: ({x1_exp}, {y1_exp}, {x2_exp}, {y2_exp})")
                x1, y1, x2, y2 = x1_exp, y1_exp, x2_exp, y2_exp

            # ========== 第2步：ResNet识别油量 ==========
            crop = img_rgb[y1:y2, x1:x2]
            crop_resized = cv2.resize(crop, (224, 224))

            # 预处理
            img_tensor = torch.from_numpy(crop_resized).float() / 255.0
            img_tensor = img_tensor.permute(2, 0, 1).unsqueeze(0).to(device)

            # 推理
            with torch.no_grad():
                fuel_pred = resnet_model(img_tensor).item()

            fuel_predictions.append(fuel_pred)
            detected_boxes.append({
                'box': (x1, y1, x2, y2),
                'conf': conf,
                'fuel': fuel_pred
            })

            print(f"   ✅ 框: 置信度 {conf:.1%}, 油量 {fuel_pred:.1%}")

        # 如果没有检测到框
        if not detected_boxes:
            print(f"   ⚠️  未检测到任何框\n")
            continue

        # ========== 绘制结果 ==========
        # 绘制检测框和油量信息（上下方都显示）
        for box_info in detected_boxes:
            x1, y1, x2, y2 = box_info['box']
            conf = box_info['conf']
            fuel = box_info['fuel']

            # 绘制框
            cv2.rectangle(img_vis, (x1, y1), (x2, y2), (0, 255, 0), 3)

            # 准备显示文字
            text_fuel = f"Fuel: {fuel:.1%}"
            text_conf = f"Conf: {conf:.1%}"
            font_scale = 0.9
            thickness = 2
            font = cv2.FONT_HERSHEY_SIMPLEX

            # 获取文字大小
            (text_w_fuel, text_h), baseline = cv2.getTextSize(text_fuel, font, font_scale, thickness)
            (text_w_conf, _), _ = cv2.getTextSize(text_conf, font, font_scale, thickness)
            text_w = max(text_w_fuel, text_w_conf)
            padding = 5

            # ===== 上方标记 =====
            # 显示两行文字：置信度 + 油量
            total_height = text_h * 2 + padding * 3
            top_bg_y1 = max(0, y1 - total_height - 5)
            top_bg_y2 = y1 - 5
            cv2.rectangle(img_vis, (x1, top_bg_y1), (x1 + text_w + 2*padding, top_bg_y2), (0, 0, 0), -1)
            # 第一行：置信度（绿色）
            cv2.putText(img_vis, text_conf, (x1 + padding, y1 - text_h - padding - 5), font, font_scale, (0, 255, 0), thickness)
            # 第二行：油量（橙色）
            cv2.putText(img_vis, text_fuel, (x1 + padding, y1 - padding - 5), font, font_scale, (0, 165, 255), thickness)

            # ===== 下方标记 =====
            # 显示两行文字：油量 + 置信度
            bottom_bg_y1 = y2 + 5
            bottom_bg_y2 = min(h, y2 + total_height + 10)
            cv2.rectangle(img_vis, (x1, bottom_bg_y1), (x1 + text_w + 2*padding, bottom_bg_y2), (0, 0, 0), -1)
            # 第一行：油量（橙色）
            cv2.putText(img_vis, text_fuel, (x1 + padding, y2 + text_h + padding + 5), font, font_scale, (0, 165, 255), thickness)
            # 第二行：置信度（绿色）
            cv2.putText(img_vis, text_conf, (x1 + padding, y2 + text_h * 2 + padding * 2 + 5), font, font_scale, (0, 255, 0), thickness)

        # 保存结果图
        result_img = cv2.cvtColor(img_vis, cv2.COLOR_RGB2BGR)
        result_path = result_dir / 'detected' / f"{img_name}_detected.jpg"
        cv2.imwrite(str(result_path), result_img)
        print(f"   💾 检测结果已保存: {result_path.name}")

        # 保存裁剪的框内图片
        for i, box_info in enumerate(detected_boxes):
            x1, y1, x2, y2 = box_info['box']
            crop = img[y1:y2, x1:x2]

            fuel = box_info['fuel']
            crop_path = result_dir / 'crops' / f"{img_name}_crop_{i+1}_fuel{fuel:.1%}.jpg"
            cv2.imwrite(str(crop_path), crop)

        # 记录结果
        if fuel_predictions:
            avg_fuel = np.mean(fuel_predictions)
            conf_values = [box['conf'] for box in detected_boxes]
            avg_conf = np.mean(conf_values)
            results_list.append({
                'image': img_name,
                'boxes_count': len(detected_boxes),
                'fuel_values': fuel_predictions,
                'avg_fuel': avg_fuel,
                'conf_values': conf_values,
                'avg_conf': avg_conf,
                'boxes': detected_boxes
            })
            print(f"   📊 平均油量: {avg_fuel:.1%}, 平均置信度: {avg_conf:.1%}\n")

    # ========== 汇总结果 ==========
    print("\n" + "="*70)
    print("【预测结果汇总】")
    print("="*70 + "\n")

    if not results_list:
        print("❌ 没有检测到任何框")
        return False

    print(f"📊 总体统计:")
    print(f"   处理图片: {len(results_list)} 张")
    print(f"   总检测框: {sum(r['boxes_count'] for r in results_list)} 个\n")

    print(f"📈 详细结果:")
    for result in results_list:
        print(f"   {result['image']}:")
        print(f"      框数: {result['boxes_count']}")
        print(f"      油量值: {[f'{f:.1%}' for f in result['fuel_values']]}")
        print(f"      平均油量: {result['avg_fuel']:.1%}")
        print(f"      置信度: {[f'{c:.1%}' for c in result['conf_values']]}")
        print(f"      平均置信度: {result['avg_conf']:.1%}")

    # ========== 保存为Excel ==========
    print(f"\n💾 导出Excel报告...")
    save_results_to_excel(results_list, result_dir)

    print(f"\n💾 结果已保存:")
    print(f"   检测图片: {result_dir / 'detected'}")
    print(f"   裁剪框图: {result_dir / 'crops'}")
    print(f"   Excel报告: {result_dir / 'prediction_results.xlsx'}\n")

    return True


def save_results_to_excel(results_list, result_dir):
    """将预测结果保存为Excel文件

    Args:
        results_list: 预测结果列表
        result_dir: 结果保存目录
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "预测结果"

    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 15

    # 写入表头
    headers = ['图片名称', '检测框数', '油量值(%)', '平均油量(%)', '置信度(%)', '平均置信度(%)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # 写入数据
    center_alignment = Alignment(horizontal="center", vertical="center")
    for row_idx, result in enumerate(results_list, 2):
        # 图片名称
        cell = ws.cell(row=row_idx, column=1)
        cell.value = result['image']
        cell.alignment = center_alignment
        cell.border = border

        # 框数
        cell = ws.cell(row=row_idx, column=2)
        cell.value = result['boxes_count']
        cell.alignment = center_alignment
        cell.border = border

        # 油量值
        fuel_values_str = ', '.join([f"{v*100:.1f}" for v in result['fuel_values']])
        cell = ws.cell(row=row_idx, column=3)
        cell.value = fuel_values_str
        cell.alignment = center_alignment
        cell.border = border

        # 平均油量
        cell = ws.cell(row=row_idx, column=4)
        cell.value = result['avg_fuel'] * 100
        cell.number_format = '0.0'
        cell.alignment = center_alignment
        cell.border = border

        # 置信度值
        conf_values_str = ', '.join([f"{c*100:.1f}" for c in result['conf_values']])
        cell = ws.cell(row=row_idx, column=5)
        cell.value = conf_values_str
        cell.alignment = center_alignment
        cell.border = border

        # 平均置信度
        cell = ws.cell(row=row_idx, column=6)
        cell.value = result['avg_conf'] * 100
        cell.number_format = '0.0'
        cell.alignment = center_alignment
        cell.border = border

    # 添加统计行
    stat_row = len(results_list) + 3
    stat_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    stat_font = Font(bold=True, size=11)

    # 统计标题
    cell = ws.cell(row=stat_row, column=1)
    cell.value = "统计汇总"
    cell.fill = stat_fill
    cell.font = stat_font
    cell.border = border

    # 总图片数
    cell = ws.cell(row=stat_row, column=2)
    cell.value = len(results_list)
    cell.fill = stat_fill
    cell.font = stat_font
    cell.alignment = center_alignment
    cell.border = border

    # 总框数
    total_boxes = sum(r['boxes_count'] for r in results_list)
    cell = ws.cell(row=stat_row, column=3)
    cell.value = total_boxes
    cell.fill = stat_fill
    cell.font = stat_font
    cell.alignment = center_alignment
    cell.border = border


    # 保存文件
    excel_path = result_dir / 'prediction_results.xlsx'
    wb.save(str(excel_path))
    print(f"   ✅ Excel文件已保存: {excel_path.name}")


if __name__ == '__main__':
    import sys

    print("\n" + "="*70)
    print("完整预测脚本 - YOLO + ResNet50")
    print("="*70)
    print(f"⏰ 预测时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

    dt = '0520'
    foldername = '0520识别错误'
    # ========== 默认值配置（可在此修改） ==========
    DEFAULT_IMAGE_DIR = 'E:/predict/'+ dt +'/'+ foldername  # 默认输入目录
    DEFAULT_RESULT_DIR = 'E:/predict/'+ dt +'/'+ foldername + 'predict'  # 默认输出目录

    # DEFAULT_IMAGE_DIR = "E:/predict/0408/0408重点网点白班打卡"  # 默认输入目录
    # DEFAULT_RESULT_DIR = "E:/predict/0408/0408重点网点白班打卡predict" # 默认输入目录

    # DEFAULT_IMAGE_DIR = "E:/predict/0408/0408重点网点夜班打卡"  # 默认输入目录
    # DEFAULT_RESULT_DIR = "E:/predict/0408/0408重点网点夜班打卡predict" # 默认输入目录

    DEFAULT_YOLO_CONF = 0.3                                      # 默认置信度
    DEFAULT_MIN_BOX_SIZE = 30                                    # 默认最小框大小（像素）
    DEFAULT_MIN_BOX_RATIO = 0.4                                  # 默认最小框比例（高/宽）
    # ============================================

    # 参数处理
    if len(sys.argv) < 2:
        print("用法: python predict_complete.py [image_path_or_dir] [result_dir] [yolo_conf] [min_box_size] [min_box_ratio]")
        print("\n参数说明:")
        print("   image_path_or_dir: 输入图片路径或目录（可选）")
        print("   result_dir: 输出结果目录（可选）")
        print("   yolo_conf: YOLO置信度阈值（可选，范围: 0.0-1.0）")
        print("   min_box_size: 最小框大小（可选，单位像素，默认80）")
        print("   min_box_ratio: 最小框比例（可选，高/宽，默认0.4，越大框越方正）")
        print("\n默认值配置:")
        print(f"   输入目录: {DEFAULT_IMAGE_DIR}")
        print(f"   输出目录: {DEFAULT_RESULT_DIR}")
        print(f"   置信度: {DEFAULT_YOLO_CONF}")
        print(f"   最小框大小: {DEFAULT_MIN_BOX_SIZE} 像素")
        print(f"   最小框比例: {DEFAULT_MIN_BOX_RATIO}")
        print("\n示例:")
        print("   python predict_complete.py                                      # 使用全部默认值")
        print("   python predict_complete.py ./test_images                        # 指定输入目录")
        print("   python predict_complete.py ./test_images ./results              # 指定输入和输出")
        print("   python predict_complete.py ./test_images ./results 0.5          # 指定置信度")
        print("   python predict_complete.py ./test_images ./results 0.5 100 0.5  # 全部指定\n")

        # 使用默认值运行
        print("⚙️  使用默认配置启动...")
        image_path = DEFAULT_IMAGE_DIR
        result_dir = DEFAULT_RESULT_DIR
        yolo_conf = DEFAULT_YOLO_CONF
        min_box_size = DEFAULT_MIN_BOX_SIZE
        min_box_ratio = DEFAULT_MIN_BOX_RATIO
    else:
        # 读取命令行参数
        image_path = sys.argv[1]
        result_dir = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_RESULT_DIR
        yolo_conf = float(sys.argv[3]) if len(sys.argv) > 3 else DEFAULT_YOLO_CONF
        min_box_size = int(sys.argv[4]) if len(sys.argv) > 4 else DEFAULT_MIN_BOX_SIZE
        min_box_ratio = float(sys.argv[5]) if len(sys.argv) > 5 else DEFAULT_MIN_BOX_RATIO

    # 参数验证
    if not (0.0 <= yolo_conf <= 1.0):
        print(f"❌ 错误：YOLO置信度必须在 0.0-1.0 之间，输入值: {yolo_conf}")
        sys.exit(1)

    if min_box_size < 10:
        print(f"❌ 错误：最小框大小必须 >= 10，输入值: {min_box_size}")
        sys.exit(1)

    if not (0.1 <= min_box_ratio <= 1.0):
        print(f"❌ 错误：最小框比例必须在 0.1-1.0 之间，输入值: {min_box_ratio}")
        sys.exit(1)

    print(f"📁 输入: {image_path}")
    print(f"💾 输出: {result_dir}")
    print(f"🎯 置信度: {yolo_conf:.1%}")
    print(f"📏 最小框大小: {min_box_size}x{min_box_size} 像素")
    print(f"📐 最小框比例(高/宽): {min_box_ratio:.1f}\n")

    predict_complete(image_path, result_dir=result_dir, yolo_conf=yolo_conf, min_box_size=min_box_size, min_box_ratio=min_box_ratio)